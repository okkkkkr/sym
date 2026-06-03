import asyncio
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Body, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from tortoise.expressions import Q

from app.controllers.brand import brand_controller
from app.controllers.category import category_controller
from app.controllers.tag import tag_controller
from app.models.admin import Product
from app.settings import settings
from app.schemas.base import DeleteIdsIn, Success, SuccessExtra
from app.schemas.categories import CategoryCreate, CategoryHotConfigUpdate, CategoryImportItem, CategoryInheritIn, CategoryUpdate
from app.utils.excel_export import build_xlsx_content

router = APIRouter()

TEMPLATE_HEADERS = ["分类名称", "分类描述", "排序", "是否启用"]


async def serialize_category_payload(category_obj):
    item = await category_obj.to_dict()
    item["product_count"] = await Product.filter(category_id=category_obj.id).count()
    return item


def parse_optional_bool(value: Any) -> bool | None:
    if value in (None, "", "all"):
        return None
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized == "true":
        return True
    if normalized == "false":
        return False
    return None


def build_category_search(name: str = "", is_active=None) -> tuple[Q, bool]:
    q = Q()
    has_filters = False
    normalized_name = str(name or "").strip()
    normalized_status = parse_optional_bool(is_active)

    if normalized_name:
        q &= Q(name__contains=normalized_name)
        has_filters = True
    if normalized_status is not None:
        q &= Q(is_active=normalized_status)
        has_filters = True

    return q, has_filters


def normalize_import_bool(value: Any, row_index: int) -> bool:
    normalized = str(value or "1").strip()
    if normalized not in {"1", "0"}:
        raise HTTPException(status_code=400, detail=f"第 {row_index} 行是否启用值不合法，仅支持 1/0")
    return normalized == "1"


async def resolve_category_ids(payload: DeleteIdsIn) -> list[int]:
    if payload.scope == "selected":
        return payload.ids

    if payload.scope == "filtered":
        search, has_filters = build_category_search(
            name=payload.filters.get("name", ""),
            is_active=payload.filters.get("is_active"),
        )
        if not has_filters:
            return []
        return list(await category_controller.model.filter(search).values_list("id", flat=True))

    return list(await category_controller.model.all().values_list("id", flat=True))


@router.get("/list", summary="查看类目列表")
async def list_category(
    page: int = Query(1, description="页码"),
    page_size: int = Query(10, description="每页数量"),
    name: str = Query("", description="类目名称"),
    is_active: bool | None = Query(None, description="是否启用"),
    sort_field: str | None = Query(None, description="排序字段"),
    sort_order: str | None = Query(None, description="排序方向 asc/desc"),
):
    q, _ = build_category_search(name=name, is_active=is_active)
    order = category_controller.build_order(
        default_order=["-updated_at", "-id"],
        sort_field=sort_field,
        sort_order=sort_order,
        allowed_fields={"updated_at", "name", "order", "is_active"},
    )
    total, category_objs = await category_controller.list(page=page, page_size=page_size, search=q, order=order)
    data = await asyncio.gather(*(serialize_category_payload(obj) for obj in category_objs))
    return SuccessExtra(data=data, total=total, page=page, page_size=page_size)


@router.get("/get", summary="查看类目")
async def get_category(id: int = Query(..., description="类目ID")):
    category_obj = await category_controller.get(id=id)
    return Success(data=await serialize_category_payload(category_obj))


@router.post("/create", summary="创建类目")
async def create_category(category_in: CategoryCreate):
    await category_controller.create(obj_in=category_in)
    return Success(msg="Created Successfully")


@router.post("/update", summary="更新类目")
async def update_category(category_in: CategoryUpdate):
    await category_controller.update(id=category_in.id, obj_in=category_in)
    return Success(msg="Updated Successfully")


@router.post("/inherit", summary="继承类目内容")
async def inherit_category_content(payload: CategoryInheritIn):
    data = await category_controller.inherit_content(source_id=payload.source_id, target_id=payload.target_id)
    return Success(msg="Inherited Successfully", data=data.model_dump())


@router.delete("/delete", summary="删除类目")
async def delete_category(payload: DeleteIdsIn = Body(...)):
    ids = await resolve_category_ids(payload)
    deleted_count = await category_controller.remove_many(ids=ids)
    return Success(msg="Deleted Successfully", data={"deleted": deleted_count})


@router.post("/export", summary="批量导出类目")
async def export_category(payload: DeleteIdsIn = Body(...)):
    ids = await resolve_category_ids(payload)
    category_objs = await category_controller.model.filter(id__in=ids).order_by("order", "-updated_at", "-id")

    rows = []
    for category_obj in category_objs:
        item = await serialize_category_payload(category_obj)
        rows.append(
            [
                item.get("name") or "",
                item.get("desc") or "",
                item.get("product_count") or 0,
                item.get("order") or 0,
                "启用" if item.get("is_active") else "停用",
                item.get("updated_at") or "",
            ]
        )

    content = build_xlsx_content(
        sheet_title="分类导出",
        headers=["分类名称", "分类描述", "关联好物数", "排序", "是否启用", "更新时间"],
        rows=rows,
    )
    filename = f'category-export-{payload.scope}-{settings.VERSION}.xlsx'
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/template", summary="下载分类导入模板")
async def download_category_template():
    content = await category_controller.build_import_template()
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="category-import-template.xlsx"'},
    )


@router.post("/import", summary="批量导入分类")
async def import_categories(file: UploadFile = File(..., description="XLSX模板文件")):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 XLSX 模板导入")

    content = await file.read()
    try:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="文件解析失败，请使用模板文件导入") from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入文件为空")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if header[: len(TEMPLATE_HEADERS)] != TEMPLATE_HEADERS:
        raise HTTPException(status_code=400, detail="模板表头不正确，请重新下载模板")

    grouped_items: dict[str, dict[str, Any]] = {}
    for index, row in enumerate(rows[1:], start=2):
        row_map = {
            TEMPLATE_HEADERS[position]: row[position] if position < len(row) else None
            for position in range(len(TEMPLATE_HEADERS))
        }
        if all(value in (None, "") for value in row_map.values()):
            continue

        name = str(row_map.get("分类名称") or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail=f"第 {index} 行分类名称不能为空")

        try:
            order = int(row_map.get("排序") or 0)
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=f"第 {index} 行排序值不合法") from exc

        normalized_item = {
            "name": name,
            "desc": str(row_map.get("分类描述") or "").strip() or None,
            "order": order,
            "is_active": normalize_import_bool(row_map.get("是否启用"), index),
        }

        existing_item = grouped_items.get(name)
        if existing_item is None:
            grouped_items[name] = normalized_item
            continue

        comparable_existing = {key: existing_item[key] for key in ["desc", "order", "is_active"]}
        comparable_current = {key: normalized_item[key] for key in ["desc", "order", "is_active"]}
        if comparable_existing != comparable_current:
            raise HTTPException(status_code=400, detail=f"第 {index} 行与分类 {name} 的其他字段不一致")

    items = [CategoryImportItem(**item) for item in grouped_items.values()]
    created = await category_controller.import_items(items)
    return Success(msg="Imported Successfully", data={"created": created})


@router.get("/hot-config", summary="查看类目热门配置")
async def get_category_hot_config(id: int = Query(..., description="类目ID")):
    category_obj = await category_controller.get(id=id)
    brands = await brand_controller.model.filter(categories__id=category_obj.id, is_active=True).distinct().order_by(
        "order", "-updated_at", "-id"
    )
    tags = await tag_controller.model.all().order_by("sort", "-updated_at", "-id")
    hot_brand_ids = [brand.id for brand in await category_obj.hot_brands.all()]
    hot_tag_ids = [tag.id for tag in await category_obj.hot_tags.all()]
    return Success(
        data={
            "id": category_obj.id,
            "hot_brand_ids": hot_brand_ids,
            "hot_tag_ids": hot_tag_ids,
            "brands": [{"id": brand.id, "name": brand.name} for brand in brands],
            "tags": [{"id": tag.id, "name": tag.name} for tag in tags],
        }
    )


@router.post("/hot-config", summary="更新类目热门配置")
async def update_category_hot_config(payload: CategoryHotConfigUpdate):
    category_obj = await category_controller.get(id=payload.id)

    normalized_brand_ids = list(dict.fromkeys(payload.hot_brand_ids))
    normalized_tag_ids = list(dict.fromkeys(payload.hot_tag_ids))

    if normalized_brand_ids:
        found_brand_count = await brand_controller.model.filter(
            id__in=normalized_brand_ids, categories__id=category_obj.id, is_active=True
        ).distinct().count()
        if found_brand_count != len(normalized_brand_ids):
            raise HTTPException(status_code=400, detail="hot_brand_ids contains invalid brand")

    if normalized_tag_ids:
        found_tag_count = await tag_controller.model.filter(id__in=normalized_tag_ids).count()
        if found_tag_count != len(normalized_tag_ids):
            raise HTTPException(status_code=400, detail="hot_tag_ids contains invalid tag")

    await category_obj.hot_brands.clear()
    if normalized_brand_ids:
        brand_objs = await brand_controller.model.filter(id__in=normalized_brand_ids)
        await category_obj.hot_brands.add(*brand_objs)

    await category_obj.hot_tags.clear()
    if normalized_tag_ids:
        tag_objs = await tag_controller.model.filter(id__in=normalized_tag_ids)
        await category_obj.hot_tags.add(*tag_objs)

    return Success(msg="Updated Successfully")
