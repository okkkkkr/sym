import asyncio
import re
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Body, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from tortoise.expressions import Q

from app.controllers.brand import brand_controller
from app.controllers.category import category_controller
from app.models.admin import Product
from app.settings import settings
from app.schemas.base import DeleteIdsIn, Success, SuccessExtra
from app.schemas.brands import BrandCreate, BrandImportItem, BrandInheritIn, BrandUpdate
from app.schemas.sortable import parse_import_rank_value
from app.utils.excel_export import build_xlsx_content

router = APIRouter()

TEMPLATE_HEADERS = ["品牌名称", "所属分类", "品牌描述", "检索次数", "排序", "是否启用"]


def parse_semicolon_values(value: Any) -> list[str]:
    if value in (None, ""):
        return []
    parts = re.split(r"[;；\n]+", str(value))
    return list(dict.fromkeys(part.strip() for part in parts if part and part.strip()))


async def ensure_category_ids_exist(category_ids: list[int]) -> list[int]:
    normalized_category_ids = list(dict.fromkeys(category_ids))
    if not normalized_category_ids:
        raise HTTPException(status_code=400, detail="category_ids must not be empty")

    found_count = await category_controller.model.filter(id__in=normalized_category_ids).count()
    if found_count != len(normalized_category_ids):
        raise HTTPException(status_code=400, detail="category_ids contains invalid category")

    return normalized_category_ids


async def serialize_brand_payload(brand_obj):
    item = await brand_obj.to_dict()
    categories = [await category.to_dict() for category in await brand_obj.categories.all()]
    item["categories"] = categories
    item["category_ids"] = [category["id"] for category in categories]
    item["category"] = categories[0] if categories else None
    item["product_count"] = await Product.filter(brand_id=brand_obj.id).count()
    return item


def parse_optional_bool(value: Any) -> bool | None:
    if value in (None, "", "all"):
        return None
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized == "true":
        return True
    if normalized == "false":
        return False
    return None


def parse_optional_int(value: Any) -> int | None:
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def build_brand_search(name: str = "", category_id: Any = None, is_active: Any = None) -> tuple[Q, bool]:
    q = Q()
    has_filters = False
    normalized_name = str(name or "").strip()
    normalized_category_id = parse_optional_int(category_id)
    normalized_status = parse_optional_bool(is_active)

    if normalized_name:
        q &= Q(name__contains=normalized_name)
        has_filters = True
    if normalized_category_id is not None:
        q &= Q(categories__id=normalized_category_id)
        has_filters = True
    if normalized_status is not None:
        q &= Q(is_active=normalized_status)
        has_filters = True

    return q, has_filters


async def resolve_brand_ids(payload: DeleteIdsIn) -> list[int]:
    if payload.scope == "selected":
        return payload.ids

    if payload.scope == "filtered":
        search, has_filters = build_brand_search(
            name=payload.filters.get("name", ""),
            category_id=payload.filters.get("category_id"),
            is_active=payload.filters.get("is_active"),
        )
        if not has_filters:
            return []
        return list(await brand_controller.model.filter(search).distinct().values_list("id", flat=True))

    return list(await brand_controller.model.all().distinct().values_list("id", flat=True))


@router.get("/list", summary="查看品牌列表")
async def list_brand(
    page: int = Query(1, description="页码"),
    page_size: int = Query(10, description="每页数量"),
    name: str = Query("", description="品牌名称"),
    category_id: int | None = Query(None, description="类目ID"),
    is_active: bool | None = Query(None, description="是否启用"),
    sort_field: str | None = Query(None, description="排序字段"),
    sort_order: str | None = Query(None, description="排序方向 asc/desc"),
):
    q, _ = build_brand_search(name=name, category_id=category_id, is_active=is_active)
    annotations = None
    order = brand_controller.build_order(
        default_order=["-updated_at", "-id"],
        sort_field=sort_field,
        sort_order=sort_order,
        allowed_fields={"updated_at", "name", "order", "search_count", "is_active"},
    )
    if sort_field == "order":
        annotations, order = brand_controller.build_nullable_field_order("order", ["-updated_at", "-id"], sort_order)
    total, brand_objs = await brand_controller.list(
        page=page,
        page_size=page_size,
        search=q,
        order=order,
        annotations=annotations,
    )
    data = await asyncio.gather(*(serialize_brand_payload(obj) for obj in brand_objs))
    return SuccessExtra(data=data, total=total, page=page, page_size=page_size)


@router.get("/get", summary="查看品牌")
async def get_brand(id: int = Query(..., description="品牌ID")):
    brand_obj = await brand_controller.get(id=id)
    data = await serialize_brand_payload(brand_obj)
    return Success(data=data)


@router.post("/create", summary="创建品牌")
async def create_brand(brand_in: BrandCreate):
    brand_in.category_ids = await ensure_category_ids_exist(brand_in.category_ids)
    await brand_controller.create_with_categories(obj_in=brand_in)
    return Success(msg="Created Successfully")


@router.post("/update", summary="更新品牌")
async def update_brand(brand_in: BrandUpdate):
    brand_in.category_ids = await ensure_category_ids_exist(brand_in.category_ids)
    await brand_controller.update_with_categories(id=brand_in.id, obj_in=brand_in)
    return Success(msg="Updated Successfully")


@router.post("/inherit", summary="继承品牌内容")
async def inherit_brand_content(payload: BrandInheritIn):
    data = await brand_controller.inherit_content(source_id=payload.source_id, target_id=payload.target_id)
    return Success(msg="Inherited Successfully", data=data.model_dump())


@router.delete("/delete", summary="删除品牌")
async def delete_brand(payload: DeleteIdsIn = Body(...)):
    ids = await resolve_brand_ids(payload)
    deleted_count = await brand_controller.remove_many(ids=ids)
    return Success(msg="Deleted Successfully", data={"deleted": deleted_count})


@router.post("/export", summary="批量导出品牌")
async def export_brand(payload: DeleteIdsIn = Body(...)):
    ids = await resolve_brand_ids(payload)
    order_annotations, export_order = brand_controller.build_nullable_field_order("order", ["-updated_at", "-id"])
    brand_objs = await brand_controller.model.filter(id__in=ids).distinct().annotate(**order_annotations).order_by(*export_order)

    rows = []
    for brand_obj in brand_objs:
        item = await serialize_brand_payload(brand_obj)
        rows.append(
            [
                item.get("name") or "",
                ";".join(category.get("name") or "" for category in item.get("categories") or []),
                item.get("desc") or "",
                item.get("product_count") or 0,
                item.get("search_count") or 0,
                "启用" if item.get("is_active") else "停用",
                item.get("updated_at") or "",
                item.get("order") if item.get("order") is not None else "",
            ]
        )

    content = build_xlsx_content(
        sheet_title="品牌导出",
        headers=["品牌名称", "所属分类", "品牌描述", "关联好物数", "检索次数", "是否启用", "更新时间", "排序"],
        rows=rows,
    )
    filename = f'brand-export-{payload.scope}-{settings.VERSION}.xlsx'
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/template", summary="下载品牌导入模板")
async def download_brand_template():
    content = await brand_controller.build_import_template()
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="brand-import-template.xlsx"'},
    )


@router.post("/import", summary="批量导入品牌")
async def import_brands(file: UploadFile = File(..., description="XLSX模板文件")):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 XLSX 模板导入")

    content = await file.read()
    try:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="文件解析失败，请使用模板文件导入") from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入文件为空")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if header[: len(TEMPLATE_HEADERS)] != TEMPLATE_HEADERS:
        raise HTTPException(status_code=400, detail="模板表头不正确，请重新下载模板")

    category_rows = await category_controller.model.all().values("id", "name")
    category_name_to_id = {str(item["name"]).strip(): item["id"] for item in category_rows if item.get("name")}

    grouped_items: dict[str, dict[str, Any]] = {}
    for index, row in enumerate(rows[1:], start=2):
        row_map = {
            TEMPLATE_HEADERS[position]: row[position] if position < len(row) else None
            for position in range(len(TEMPLATE_HEADERS))
        }
        if all(value in (None, "") for value in row_map.values()):
            continue

        name = str(row_map.get("品牌名称") or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail=f"第 {index} 行品牌名称不能为空")

        category_names = parse_semicolon_values(row_map.get("所属分类"))
        if not category_names:
            raise HTTPException(status_code=400, detail=f"第 {index} 行所属分类不能为空")
        invalid_category_names = [category_name for category_name in category_names if category_name not in category_name_to_id]
        if invalid_category_names:
            raise HTTPException(
                status_code=400,
                detail=f"第 {index} 行所属分类不存在: {';'.join(invalid_category_names)}",
            )

        is_active_value = str(row_map.get("是否启用") or "1").strip()
        if is_active_value not in {"1", "0"}:
            raise HTTPException(status_code=400, detail=f"第 {index} 行是否启用值不合法，仅支持 1/0")

        try:
            search_count = int(row_map.get("检索次数") or 0)
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=f"第 {index} 行检索次数不合法") from exc
        try:
            order = parse_import_rank_value(row_map.get("排序"))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"第 {index} 行{exc}") from exc

        normalized_item = {
            "name": name,
            "desc": str(row_map.get("品牌描述") or "").strip() or None,
            "search_count": search_count,
            "order": order,
            "is_active": is_active_value == "1",
        }

        existing_item = grouped_items.get(name)
        if existing_item is None:
            grouped_items[name] = {
                **normalized_item,
                "category_ids": [category_name_to_id[category_name] for category_name in category_names],
            }
            continue

        comparable_existing = {
            key: existing_item[key]
            for key in ["desc", "search_count", "order", "is_active"]
        }
        comparable_current = {
            key: normalized_item[key]
            for key in ["desc", "search_count", "order", "is_active"]
        }
        if comparable_existing != comparable_current:
            raise HTTPException(status_code=400, detail=f"第 {index} 行与品牌 {name} 的其他字段不一致")

        for category_name in category_names:
            category_id = category_name_to_id[category_name]
            if category_id not in existing_item["category_ids"]:
                existing_item["category_ids"].append(category_id)

    items = [BrandImportItem(**item) for item in grouped_items.values()]

    created = await brand_controller.import_items(items)
    return Success(msg="Imported Successfully", data={"created": created})
