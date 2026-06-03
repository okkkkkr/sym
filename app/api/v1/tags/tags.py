from io import BytesIO
from typing import Any

from fastapi import APIRouter, Body, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from tortoise.expressions import Q
from tortoise.functions import Count

from app.controllers.tag import tag_controller
from app.schemas.base import DeleteIdsIn, Success, SuccessExtra
from app.schemas.tags import TagCreate, TagImportItem, TagUpdate
from app.settings import settings
from app.utils.excel_export import build_xlsx_content

router = APIRouter()

TEMPLATE_HEADERS = ["标签名称", "备注", "检索次数", "排序", "是否启用"]


async def serialize_tag_payload(tag_obj) -> dict:
    return {
        "id": tag_obj.id,
        "name": tag_obj.name,
        "remark": tag_obj.remark,
        "search_count": tag_obj.search_count,
        "sort": tag_obj.sort,
        "is_active": tag_obj.is_active,
        "product_count": getattr(tag_obj, "product_count", 0),
        "created_at": tag_obj.created_at.strftime(settings.DATETIME_FORMAT) if tag_obj.created_at else None,
        "updated_at": tag_obj.updated_at.strftime(settings.DATETIME_FORMAT) if tag_obj.updated_at else None,
    }


def parse_optional_bool(value: Any) -> bool | None:
    if value in (None, "", "all"):
        return None
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized == "true":
        return True
    if normalized == "false":
        return False
    return None


def build_tag_search(name: str = "", is_active: Any = None) -> tuple[Q, bool]:
    q = Q()
    has_filters = False
    normalized_name = str(name or "").strip()
    normalized_status = parse_optional_bool(is_active)

    if normalized_name:
        q &= Q(name__contains=normalized_name)
        has_filters = True
    if normalized_status is not None:
        q &= Q(is_active=normalized_status)
        has_filters = True

    return q, has_filters


async def resolve_tag_ids(payload: DeleteIdsIn) -> list[int]:
    if payload.scope == "selected":
        return payload.ids

    if payload.scope == "filtered":
        search, has_filters = build_tag_search(
            name=payload.filters.get("name", ""),
            is_active=payload.filters.get("is_active"),
        )
        if not has_filters:
            return []
        return list(await tag_controller.model.filter(search).values_list("id", flat=True))

    return list(await tag_controller.model.all().values_list("id", flat=True))


@router.get("/list", summary="查看标签列表")
async def list_tag(
    page: int = Query(1, description="页码"),
    page_size: int = Query(10, description="每页数量"),
    name: str = Query("", description="标签名称"),
    is_active: bool | None = Query(None, description="是否启用"),
    sort_field: str | None = Query(None, description="排序字段"),
    sort_order: str | None = Query(None, description="排序方向 asc/desc"),
):
    q, _ = build_tag_search(name=name, is_active=is_active)
    order = tag_controller.build_order(
        default_order=["-updated_at", "-id"],
        sort_field=sort_field,
        sort_order=sort_order,
        allowed_fields={"updated_at", "name", "search_count", "sort", "is_active", "product_count"},
    )
    total, tag_objs = await tag_controller.list_with_product_count(page=page, page_size=page_size, search=q, order=order)
    data = [await serialize_tag_payload(obj) for obj in tag_objs]
    return SuccessExtra(data=data, total=total, page=page, page_size=page_size)


@router.get("/get", summary="查看标签")
async def get_tag(id: int = Query(..., description="标签ID")):
    tag_obj = await tag_controller.model.annotate(product_count=Count("products", distinct=True)).get(id=id)
    return Success(data=await serialize_tag_payload(tag_obj))


@router.post("/create", summary="创建标签")
async def create_tag(tag_in: TagCreate):
    await tag_controller.create(obj_in=tag_in)
    return Success(msg="Created Successfully")


@router.post("/update", summary="更新标签")
async def update_tag(tag_in: TagUpdate):
    await tag_controller.update(id=tag_in.id, obj_in=tag_in)
    return Success(msg="Updated Successfully")


@router.post("/toggle", summary="切换标签启用状态")
async def toggle_tag(id: int = Query(..., description="标签ID"), is_active: bool = Query(..., description="是否启用")):
    await tag_controller.update(id=id, obj_in={"is_active": is_active})
    return Success(msg="Updated Successfully")


@router.delete("/delete", summary="删除标签")
async def delete_tag(payload: DeleteIdsIn = Body(...)):
    ids = await resolve_tag_ids(payload)
    deleted_count = await tag_controller.remove_many(ids=ids)
    return Success(msg="Deleted Successfully", data={"deleted": deleted_count})


@router.post("/export", summary="批量导出标签")
async def export_tag(payload: DeleteIdsIn = Body(...)):
    ids = await resolve_tag_ids(payload)
    tag_objs = await tag_controller.model.filter(id__in=ids).annotate(product_count=Count("products", distinct=True)).order_by(
        "sort", "-updated_at", "-id"
    )

    rows = []
    for tag_obj in tag_objs:
        item = await serialize_tag_payload(tag_obj)
        rows.append(
            [
                item.get("name") or "",
                item.get("remark") or "",
                item.get("search_count") or 0,
                item.get("sort") or 0,
                item.get("product_count") or 0,
                "启用" if item.get("is_active") else "停用",
                item.get("updated_at") or "",
            ]
        )

    content = build_xlsx_content(
        sheet_title="标签导出",
        headers=["标签名称", "备注", "检索次数", "排序", "关联好物数", "是否启用", "更新时间"],
        rows=rows,
    )
    filename = f'tag-export-{payload.scope}-{settings.VERSION}.xlsx'
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/template", summary="下载标签导入模板")
async def download_tag_template():
    content = await tag_controller.build_import_template()
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="tag-import-template.xlsx"'},
    )


@router.post("/import", summary="批量导入标签")
async def import_tags(file: UploadFile = File(..., description="XLSX模板文件")):
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 XLSX 模板导入")

    content = await file.read()
    try:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="文件解析失败，请使用模板文件导入") from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="导入文件为空")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if header[: len(TEMPLATE_HEADERS)] != TEMPLATE_HEADERS:
        raise HTTPException(status_code=400, detail="模板表头不正确，请重新下载模板")

    grouped_items: dict[str, dict[str, Any]] = {}
    for index, row in enumerate(rows[1:], start=2):
        row_map = {
            TEMPLATE_HEADERS[position]: row[position] if position < len(row) else None
            for position in range(len(TEMPLATE_HEADERS))
        }
        if all(value in (None, "") for value in row_map.values()):
            continue

        name = str(row_map.get("标签名称") or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail=f"第 {index} 行标签名称不能为空")

        is_active_value = str(row_map.get("是否启用") or "1").strip()
        if is_active_value not in {"1", "0"}:
            raise HTTPException(status_code=400, detail=f"第 {index} 行是否启用值不合法，仅支持 1/0")

        try:
            search_count = int(row_map.get("检索次数") or 0)
            sort = int(row_map.get("排序") or 0)
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=f"第 {index} 行检索次数或排序值不合法") from exc

        normalized_item = {
            "name": name,
            "remark": str(row_map.get("备注") or "").strip() or None,
            "search_count": search_count,
            "sort": sort,
            "is_active": is_active_value == "1",
        }

        existing_item = grouped_items.get(name)
        if existing_item is None:
            grouped_items[name] = normalized_item
            continue

        comparable_existing = {key: existing_item[key] for key in ["remark", "search_count", "sort", "is_active"]}
        comparable_current = {
            key: normalized_item[key]
            for key in ["remark", "search_count", "sort", "is_active"]
        }
        if comparable_existing != comparable_current:
            raise HTTPException(status_code=400, detail=f"第 {index} 行与标签 {name} 的其他字段不一致")

    items = [TagImportItem(**item) for item in grouped_items.values()]

    created = await tag_controller.import_items(items)
    return Success(msg="Imported Successfully", data={"created": created})
