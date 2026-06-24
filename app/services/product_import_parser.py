import json
import re

from fastapi import HTTPException
from openpyxl import load_workbook

from app.schemas.product_import import ProductImportParseResult, ProductImportParsedRow
from app.schemas.sortable import parse_import_rank_value


class ProductImportParserService:
    EXPECTED_HEADERS = [
        "name",
        "material_dir",
        "category_name",
        "brand_name",
        "desc",
        "tag_names",
        "product_code_custom",
        "status",
        "order",
        "detail_text",
        "detail_description_json",
    ]
    DISPLAY_HEADERS = {
        "name": "名称",
        "material_dir": "素材目录",
        "category_name": "所属分类",
        "brand_name": "所属品牌",
        "desc": "简介",
        "tag_names": "标签",
        "product_code_custom": "好物识别码",
        "status": "上架状态",
        "order": "排序",
        "detail_text": "详情文本",
        "detail_description_json": "结构化详情JSON",
    }
    HEADER_ALIASES = {key: [key, label] for key, label in DISPLAY_HEADERS.items()}
    HEADER_LOOKUP = {alias: key for key, aliases in HEADER_ALIASES.items() for alias in aliases}
    REQUIRED_HEADERS = ["name", "material_dir", "category_name", "brand_name"]
    STATUS_TRUE_VALUES = {"true", "1", "是", "yes"}
    STATUS_FALSE_VALUES = {"false", "0", "否", "no"}

    async def parse(self, workbook_path: str) -> ProductImportParseResult:
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            headers: list[str] = []
            parsed_rows: list[ProductImportParsedRow] = []
            valid_sheet_count = 0

            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                rows = list(worksheet.iter_rows(values_only=True))
                if not rows or all(self._is_empty_row(values) for values in rows):
                    continue

                sheet_headers = [self._normalize_cell_value(value) for value in rows[0]]
                header_index = {}
                for index, header in enumerate(sheet_headers):
                    if not header:
                        continue
                    header_index[self.HEADER_LOOKUP.get(header, header)] = index
                missing_headers = [header for header in self.REQUIRED_HEADERS if header not in header_index]
                if missing_headers:
                    missing_labels = [self.DISPLAY_HEADERS.get(header, header) for header in missing_headers]
                    raise HTTPException(
                        status_code=400,
                        detail=f"Sheet「{sheet_name}」缺少必填表头: {', '.join(missing_labels)}",
                    )

                if not headers:
                    headers = sheet_headers
                valid_sheet_count += 1

                for row_no, values in enumerate(rows[1:], start=2):
                    if self._is_empty_row(values):
                        continue

                    parsed_rows.append(
                        self._build_row(
                            sheet_name=sheet_name,
                            row_no=row_no,
                            values=values,
                            header_index=header_index,
                        )
                    )

            if not valid_sheet_count:
                return ProductImportParseResult(headers=[], rows=[], total_rows=0, valid_rows=0, invalid_rows=0)

            return ProductImportParseResult(
                headers=headers,
                rows=parsed_rows,
                total_rows=len(parsed_rows),
                valid_rows=len(parsed_rows),
                invalid_rows=0,
            )
        finally:
            workbook.close()

    def _build_row(
        self, *, sheet_name: str, row_no: int, values, header_index: dict[str, int]
    ) -> ProductImportParsedRow:
        name = self._get_cell(values, header_index, "name")
        material_dir = self._get_cell(values, header_index, "material_dir")
        category_name = self._get_cell(values, header_index, "category_name")
        brand_name = self._get_cell(values, header_index, "brand_name")
        detail_text = self._get_cell(values, header_index, "detail_text")
        detail_description_json = self._get_cell(values, header_index, "detail_description_json")

        status, status_error = self._parse_status(self._get_cell(values, header_index, "status"))
        order, order_error = self._parse_order(self._get_cell(values, header_index, "order"))
        detail_description, detail_error = self._parse_detail_description(detail_text, detail_description_json)

        row = ProductImportParsedRow(
            sheet_name=sheet_name,
            row_no=row_no,
            row_label=self._build_row_label(sheet_name, row_no),
            name=name,
            material_dir=material_dir,
            category_name=category_name,
            brand_name=brand_name,
            desc=self._empty_to_none(self._get_cell(values, header_index, "desc")),
            tag_names=self._parse_tag_names(self._get_cell(values, header_index, "tag_names")),
            product_code_custom=self._empty_to_none(self._get_cell(values, header_index, "product_code_custom")),
            status=status,
            order=order,
            detail_description=detail_description,
        )
        if status_error:
            row.errors.append(status_error)
        if order_error:
            row.errors.append(order_error)
        if detail_error:
            row.errors.append(detail_error)
        return row

    @staticmethod
    def _build_row_label(sheet_name: str, row_no: int) -> str:
        return f"{sheet_name} 第{row_no}行" if sheet_name else f"第{row_no}行"

    @staticmethod
    def _normalize_cell_value(value) -> str:
        if value is None:
            return ""
        return str(value).strip()

    def _get_cell(self, values, header_index: dict[str, int], key: str) -> str:
        index = header_index.get(key)
        if index is None or index >= len(values):
            return ""
        return self._normalize_cell_value(values[index])

    @staticmethod
    def _is_empty_row(values) -> bool:
        return all(value in (None, "") for value in values)

    @staticmethod
    def _empty_to_none(value: str) -> str | None:
        return value or None

    def _parse_status(self, value: str) -> tuple[bool, str | None]:
        normalized = str(value or "").strip().lower()
        if not normalized:
            return True, None
        if normalized in self.STATUS_TRUE_VALUES:
            return True, None
        if normalized in self.STATUS_FALSE_VALUES:
            return False, None
        return True, "上架状态仅支持 true/false/1/0/是/否"

    @staticmethod
    def _parse_order(value: str) -> tuple[int | None, str | None]:
        if not value:
            return None, None
        try:
            return parse_import_rank_value(value), None
        except ValueError as exc:
            if "整数" in str(exc):
                return None, None
            return None, str(exc)

    def _parse_detail_description(self, detail_text: str, detail_description_json: str):
        if detail_description_json:
            try:
                parsed = json.loads(detail_description_json)
                if isinstance(parsed, list):
                    return parsed, None
                return [], "结构化详情JSON必须是数组"
            except json.JSONDecodeError:
                return [], "结构化详情JSON不是合法的 JSON"
        if detail_text:
            return [
                {
                    "type": "text",
                    "title": "产品介绍",
                    "content": detail_text,
                }
            ], None
        return [], None

    @staticmethod
    def _parse_tag_names(value: str) -> list[str]:
        if not value:
            return []
        parts = re.split(r"[;；\n]+", value)
        return list(dict.fromkeys(part.strip() for part in parts if part and part.strip()))


product_import_parser_service = ProductImportParserService()
